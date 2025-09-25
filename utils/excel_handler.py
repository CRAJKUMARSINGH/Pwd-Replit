"""
Excel handling utilities for PWD Tools
Read, write, and process Excel files for various PWD operations
"""

import pandas as pd
import io
from datetime import datetime
import os

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

class ExcelHandler:
    """Excel file handling utility class"""
    
    def __init__(self):
        """Initialize Excel handler"""
        pass
    
    @staticmethod
    def read_excel_file(file_path_or_buffer, sheet_name=0):
        """Read Excel file and return DataFrame"""
        try:
            if isinstance(file_path_or_buffer, str):
                # File path
                df = pd.read_excel(file_path_or_buffer, sheet_name=sheet_name)
            else:
                # File buffer (from Streamlit file uploader)
                df = pd.read_excel(file_path_or_buffer, sheet_name=sheet_name)
            
            return df
        except Exception as e:
            raise Exception(f"Error reading Excel file: {str(e)}")
    
    @staticmethod
    def write_excel_file(data, file_path=None, sheet_name='Sheet1'):
        """Write DataFrame to Excel file"""
        if file_path is None:
            # Return as buffer
            buffer = io.BytesIO()
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                if isinstance(data, dict):
                    for sheet, df in data.items():
                        df.to_excel(writer, sheet_name=sheet, index=False)
                else:
                    data.to_excel(writer, sheet_name=sheet_name, index=False)
            
            buffer.seek(0)
            return buffer.getvalue()
        else:
            # Write to file
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                if isinstance(data, dict):
                    for sheet, df in data.items():
                        df.to_excel(writer, sheet_name=sheet, index=False)
                else:
                    data.to_excel(writer, sheet_name=sheet_name, index=False)
            
            return file_path
    
    @staticmethod
    def create_formatted_excel(data, file_path=None, title="PWD Report"):
        """Create formatted Excel file with PWD styling"""
        if not OPENPYXL_AVAILABLE:
            # Fallback to basic Excel without formatting
            return ExcelHandler.write_excel_file(data, file_path)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Report"
        
        # Define styles
        header_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='004E89', end_color='004E89', fill_type='solid')
        
        subheader_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        subheader_fill = PatternFill(start_color='FF6B35', end_color='FF6B35', fill_type='solid')
        
        normal_font = Font(name='Arial', size=10)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add title
        ws['A1'] = title
        ws['A1'].font = Font(name='Arial', size=16, bold=True, color='004E89')
        ws.merge_cells('A1:E1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Add generation date
        ws['A2'] = f"Generated on: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        ws['A2'].font = Font(name='Arial', size=10, italic=True)
        ws.merge_cells('A2:E2')
        
        # Add data
        if isinstance(data, pd.DataFrame):
            # Add DataFrame
            start_row = 4
            
            # Add headers
            for col, column_name in enumerate(data.columns, 1):
                cell = ws.cell(row=start_row, column=col, value=column_name)
                cell.font = subheader_font
                cell.fill = subheader_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            
            # Add data rows
            for row_idx, row_data in enumerate(data.iterrows(), start_row + 1):
                for col_idx, value in enumerate(row_data[1], 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = normal_font
                    cell.border = border
                    
                    # Format numbers
                    if isinstance(value, (int, float)) and col_idx > 1:
                        if value > 1000:
                            cell.number_format = '#,##0.00'
                        else:
                            cell.number_format = '0.00'
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save or return buffer
        if file_path is None:
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer.getvalue()
        else:
            wb.save(file_path)
            return file_path
    
    @staticmethod
    def create_bill_excel(bill_data, file_path=None):
        """Create formatted Excel file for bill"""
        # Prepare bill items data
        if 'items' in bill_data and bill_data['items']:
            items_df = pd.DataFrame(bill_data['items'])
            items_df.index += 1
            items_df.index.name = 'S.No.'
        else:
            items_df = pd.DataFrame()
        
        # Prepare deductions data
        if 'deductions' in bill_data and bill_data['deductions']:
            deductions_df = pd.DataFrame(bill_data['deductions'])
            deductions_df.index += 1
            deductions_df.index.name = 'S.No.'
        else:
            deductions_df = pd.DataFrame()
        
        # Create workbook
        if not OPENPYXL_AVAILABLE:
            # Basic version without formatting
            sheets_data = {}
            
            # Bill summary
            summary_data = {
                'Field': ['Bill Number', 'Bill Date', 'Contractor Name', 'Project Name', 
                         'Bill Amount', 'Total Deductions', 'Net Amount'],
                'Value': [
                    bill_data.get('bill_number', ''),
                    bill_data.get('bill_date', ''),
                    bill_data.get('contractor_name', ''),
                    bill_data.get('project_name', ''),
                    f"₹{bill_data.get('bill_amount', 0):,.2f}",
                    f"₹{sum([d.get('amount', 0) for d in bill_data.get('deductions', [])]):,.2f}",
                    f"₹{bill_data.get('bill_amount', 0) - sum([d.get('amount', 0) for d in bill_data.get('deductions', [])]):,.2f}"
                ]
            }
            sheets_data['Bill Summary'] = pd.DataFrame(summary_data)
            
            if not items_df.empty:
                sheets_data['Bill Items'] = items_df
            
            if not deductions_df.empty:
                sheets_data['Deductions'] = deductions_df
            
            return ExcelHandler.write_excel_file(sheets_data, file_path)
        
        # Formatted version with openpyxl
        wb = Workbook()
        
        # Bill Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "Bill Summary"
        
        # Add bill details
        ExcelHandler._add_bill_summary(ws_summary, bill_data)
        
        # Add items sheet if items exist
        if not items_df.empty:
            ws_items = wb.create_sheet("Bill Items")
            ExcelHandler._add_items_sheet(ws_items, items_df)
        
        # Add deductions sheet if deductions exist
        if not deductions_df.empty:
            ws_deductions = wb.create_sheet("Deductions")
            ExcelHandler._add_deductions_sheet(ws_deductions, deductions_df)
        
        # Save or return buffer
        if file_path is None:
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer.getvalue()
        else:
            wb.save(file_path)
            return file_path
    
    @staticmethod
    def _add_bill_summary(ws, bill_data):
        """Add bill summary to worksheet"""
        if not OPENPYXL_AVAILABLE:
            return
        
        # Styles
        header_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='004E89', end_color='004E89', fill_type='solid')
        normal_font = Font(name='Arial', size=11)
        
        # Title
        ws['A1'] = "PWD BILL SUMMARY"
        ws['A1'].font = Font(name='Arial', size=16, bold=True, color='004E89')
        ws.merge_cells('A1:B1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Bill details
        details = [
            ('Bill Number:', bill_data.get('bill_number', 'N/A')),
            ('Bill Date:', bill_data.get('bill_date', 'N/A')),
            ('Contractor Name:', bill_data.get('contractor_name', 'N/A')),
            ('Project Name:', bill_data.get('project_name', 'N/A')),
            ('Work Order No.:', bill_data.get('work_order_no', 'N/A')),
            ('Agreement Amount:', f"₹{bill_data.get('agreement_amount', 0):,.2f}"),
            ('Bill Amount:', f"₹{bill_data.get('bill_amount', 0):,.2f}"),
            ('Total Deductions:', f"₹{sum([d.get('amount', 0) for d in bill_data.get('deductions', [])]):,.2f}"),
            ('Net Payable Amount:', f"₹{bill_data.get('bill_amount', 0) - sum([d.get('amount', 0) for d in bill_data.get('deductions', [])]):,.2f}")
        ]
        
        for idx, (label, value) in enumerate(details, 3):
            ws[f'A{idx}'] = label
            ws[f'B{idx}'] = value
            ws[f'A{idx}'].font = Font(name='Arial', size=11, bold=True)
            ws[f'B{idx}'].font = normal_font
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
    
    @staticmethod
    def _add_items_sheet(ws, items_df):
        """Add items to worksheet"""
        if not OPENPYXL_AVAILABLE:
            return
        
        # Add title
        ws['A1'] = "BILL ITEMS"
        ws['A1'].font = Font(name='Arial', size=14, bold=True, color='004E89')
        ws.merge_cells('A1:F1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Add DataFrame data starting from row 3
        for r_idx, row in enumerate(dataframe_to_rows(items_df, index=True, header=True), 3):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        # Format header row
        header_row = 4
        for col in range(1, len(items_df.columns) + 2):
            cell = ws.cell(row=header_row, column=col)
            cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='FF6B35', end_color='FF6B35', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
    
    @staticmethod
    def _add_deductions_sheet(ws, deductions_df):
        """Add deductions to worksheet"""
        if not OPENPYXL_AVAILABLE:
            return
        
        # Add title
        ws['A1'] = "DEDUCTIONS"
        ws['A1'].font = Font(name='Arial', size=14, bold=True, color='004E89')
        ws.merge_cells('A1:D1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Add DataFrame data starting from row 3
        for r_idx, row in enumerate(dataframe_to_rows(deductions_df, index=True, header=True), 3):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        # Format header row
        header_row = 4
        for col in range(1, len(deductions_df.columns) + 2):
            cell = ws.cell(row=header_row, column=col)
            cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='1A8A16', end_color='1A8A16', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
    
    @staticmethod
    def process_emd_excel(file_path_or_buffer):
        """Process EMD data from Excel file"""
        try:
            df = ExcelHandler.read_excel_file(file_path_or_buffer)
            
            # Expected columns for EMD processing
            expected_columns = [
                'tender_number', 'contractor_name', 'emd_amount', 
                'deposit_date', 'status'
            ]
            
            # Validate columns
            missing_columns = []
            for col in expected_columns:
                if col not in df.columns.str.lower().str.replace(' ', '_'):
                    missing_columns.append(col)
            
            if missing_columns:
                return None, f"Missing columns: {', '.join(missing_columns)}"
            
            # Clean and validate data
            processed_df = df.copy()
            
            # Convert date columns
            date_columns = ['deposit_date']
            for col in date_columns:
                if col in processed_df.columns:
                    processed_df[col] = pd.to_datetime(processed_df[col], errors='coerce')
            
            # Convert numeric columns
            numeric_columns = ['emd_amount']
            for col in numeric_columns:
                if col in processed_df.columns:
                    processed_df[col] = pd.to_numeric(processed_df[col], errors='coerce')
            
            # Remove rows with critical missing data
            processed_df = processed_df.dropna(subset=['tender_number', 'emd_amount'])
            
            return processed_df, None
            
        except Exception as e:
            return None, f"Error processing EMD Excel file: {str(e)}"
    
    @staticmethod
    def create_emd_template():
        """Create EMD processing template Excel file"""
        template_data = {
            'Tender Number': ['T001/2024', 'T002/2024'],
            'Contractor/Bidder Name': ['ABC Construction', 'XYZ Builders'],
            'EMD Amount': [50000, 75000],
            'Deposit Date': ['2024-01-15', '2024-01-20'],
            'Tender Date': ['2024-01-10', '2024-01-15'],
            'Work Description': ['Road Construction', 'Building Construction'],
            'Contact Number': ['9876543210', '9876543211'],
            'Bank Details': ['SBI Account', 'HDFC Account'],
            'Status': ['Active', 'Active'],
            'Remarks': ['Sample data', 'Sample data']
        }
        
        template_df = pd.DataFrame(template_data)
        
        # Create instructions
        instructions = pd.DataFrame({
            'EMD Processing Template Instructions': [
                'This template is for processing EMD (Earnest Money Deposit) data',
                '',
                'Required Columns:',
                '- Tender Number: Unique identifier for each tender',
                '- Contractor/Bidder Name: Name of the bidding organization',
                '- EMD Amount: Amount in rupees (numeric)',
                '- Deposit Date: Date in YYYY-MM-DD format',
                '',
                'Optional Columns:',
                '- All other columns as per requirement',
                '',
                'Notes:',
                '- Do not modify the header row',
                '- Ensure dates are in proper format',
                '- EMD amounts should be numeric without currency symbols',
                '',
                f'Template created: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
            ]
        })
        
        # Combine data
        sheets_data = {
            'EMD_Data': template_df,
            'Instructions': instructions
        }
        
        return ExcelHandler.write_excel_file(sheets_data)
    
    @staticmethod
    def validate_excel_structure(df, required_columns):
        """Validate Excel file structure"""
        errors = []
        warnings = []
        
        # Check if DataFrame is empty
        if df.empty:
            errors.append("Excel file is empty")
            return errors, warnings
        
        # Check for required columns
        df_columns = df.columns.str.lower().str.replace(' ', '_')
        required_lower = [col.lower().replace(' ', '_') for col in required_columns]
        
        missing_columns = []
        for req_col in required_lower:
            if req_col not in df_columns:
                missing_columns.append(req_col)
        
        if missing_columns:
            errors.append(f"Missing required columns: {', '.join(missing_columns)}")
        
        # Check for duplicate headers
        if len(df.columns) != len(set(df.columns)):
            warnings.append("Duplicate column headers detected")
        
        # Check for completely empty columns
        empty_columns = df.columns[df.isnull().all()].tolist()
        if empty_columns:
            warnings.append(f"Empty columns detected: {', '.join(empty_columns)}")
        
        # Check data types and formats
        for col in df.columns:
            col_lower = col.lower().replace(' ', '_')
            
            # Check date columns
            if 'date' in col_lower:
                try:
                    pd.to_datetime(df[col], errors='raise')
                except:
                    warnings.append(f"Column '{col}' may contain invalid dates")
            
            # Check amount columns
            if 'amount' in col_lower or 'value' in col_lower:
                try:
                    pd.to_numeric(df[col], errors='raise')
                except:
                    warnings.append(f"Column '{col}' may contain non-numeric values")
        
        return errors, warnings
    
    @staticmethod
    def clean_excel_data(df):
        """Clean Excel data for processing"""
        cleaned_df = df.copy()
        
        # Remove completely empty rows
        cleaned_df = cleaned_df.dropna(how='all')
        
        # Remove completely empty columns
        cleaned_df = cleaned_df.dropna(axis=1, how='all')
        
        # Strip whitespace from string columns
        for col in cleaned_df.select_dtypes(include=['object']).columns:
            cleaned_df[col] = cleaned_df[col].astype(str).str.strip()
            # Replace 'nan' strings with actual NaN
            cleaned_df[col] = cleaned_df[col].replace('nan', pd.NA)
        
        # Convert date-like columns
        date_keywords = ['date', 'time']
        for col in cleaned_df.columns:
            if any(keyword in col.lower() for keyword in date_keywords):
                try:
                    cleaned_df[col] = pd.to_datetime(cleaned_df[col], errors='coerce')
                except:
                    pass
        
        # Convert numeric-like columns
        numeric_keywords = ['amount', 'value', 'rate', 'percent', 'qty', 'quantity']
        for col in cleaned_df.columns:
            if any(keyword in col.lower() for keyword in numeric_keywords):
                try:
                    # Remove currency symbols and commas
                    if cleaned_df[col].dtype == 'object':
                        cleaned_df[col] = cleaned_df[col].astype(str).str.replace('[₹,\$]', '', regex=True)
                    cleaned_df[col] = pd.to_numeric(cleaned_df[col], errors='coerce')
                except:
                    pass
        
        return cleaned_df

# Utility functions
def create_excel_from_dict(data_dict, filename=None):
    """Create Excel file from dictionary of DataFrames"""
    return ExcelHandler.write_excel_file(data_dict, filename)

def read_excel_sheets(file_path_or_buffer):
    """Read all sheets from Excel file"""
    try:
        excel_file = pd.ExcelFile(file_path_or_buffer)
        sheets_data = {}
        
        for sheet_name in excel_file.sheet_names:
            sheets_data[sheet_name] = pd.read_excel(excel_file, sheet_name=sheet_name)
        
        return sheets_data, None
    except Exception as e:
        return None, f"Error reading Excel sheets: {str(e)}"

def get_excel_info(file_path_or_buffer):
    """Get information about Excel file"""
    try:
        if hasattr(file_path_or_buffer, 'read'):
            # File buffer
            file_size = len(file_path_or_buffer.getvalue())
            excel_file = pd.ExcelFile(file_path_or_buffer)
        else:
            # File path
            file_size = os.path.getsize(file_path_or_buffer)
            excel_file = pd.ExcelFile(file_path_or_buffer)
        
        info = {
            'file_size': file_size,
            'sheet_names': excel_file.sheet_names,
            'sheet_count': len(excel_file.sheet_names),
            'sheets_info': {}
        }
        
        # Get info for each sheet
        for sheet_name in excel_file.sheet_names:
            df = pd.read_excel(excel_file, sheet_name=sheet_name)
            info['sheets_info'][sheet_name] = {
                'rows': len(df),
                'columns': len(df.columns),
                'column_names': list(df.columns)
            }
        
        return info, None
    except Exception as e:
        return None, f"Error getting Excel file info: {str(e)}"
